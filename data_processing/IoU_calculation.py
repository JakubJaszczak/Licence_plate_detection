import os
from enum import Enum
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def yolo_to_corners(x_c, y_c, w, h):
    x_min = x_c - (w / 2)
    y_min = y_c - (h / 2)
    x_max = x_c + (w / 2)
    y_max = y_c + (h / 2)
    return [x_min, y_min, x_max, y_max]


def calculate_iou(box1, box2):
    x_left = max(box1[0], box2[0])
    y_top = max(box1[1], box2[1])
    x_right = min(box1[2], box2[2])
    y_bottom = min(box1[3], box2[3])

    if x_right < x_left or y_bottom < y_top:
        return 0.0

    intersection_area = (x_right - x_left) * (y_bottom - y_top)
    box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
    box2_area = (box2[2] - box2[0]) * (box2[3] - box2[1])

    iou = intersection_area / float(box1_area + box2_area - intersection_area)
    return iou


def read_gt_file(filepath):
    boxes = []
    if not os.path.exists(filepath):
        return boxes

    with open(filepath, "r") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 5:
                class_id = int(parts[0])
                x_c, y_c, w, h = map(float, parts[1:5])
                corners = yolo_to_corners(x_c, y_c, w, h)
                boxes.append({"class_id": class_id, "corners": corners, "raw": [x_c, y_c, w, h]})
    return boxes


def read_pred_file(filepath):
    boxes = []
    if not os.path.exists(filepath):
        return boxes

    with open(filepath, "r") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 5:
                class_id = int(parts[0])
                x1, y1, x2, y2 = map(float, parts[1:5])
                if "yolo" in filepath.lower():
                    x1, y1, x2, y2 = yolo_to_corners(x1, y1, x2, y2)
                confidence = float(parts[5]) if len(parts) >= 6 else None
                boxes.append(
                    {
                        "class_id": class_id,
                        "corners": [x1, y1, x2, y2],
                        "raw": [x1, y1, x2, y2],
                        "confidence": confidence,
                    }
                )
    return boxes


def process_dataset(labels_dir, preds_dir, iou_threshold=0.5):
    results = []

    for filename in os.listdir(labels_dir):
        if not filename.endswith(".txt"):
            continue

        image_name = filename.replace(".txt", "")
        gt_path = os.path.join(labels_dir, filename)
        pred_path = os.path.join(preds_dir, image_name, "results.txt")

        gt_boxes = read_gt_file(gt_path)
        pred_boxes = read_pred_file(pred_path)

        matched_pred_indices = set()

        for gt_idx, gt in enumerate(gt_boxes):
            best_iou = 0.0
            best_pred_idx = -1

            for pred_idx, pred in enumerate(pred_boxes):
                if pred_idx in matched_pred_indices:
                    continue

                iou = calculate_iou(gt["corners"], pred["corners"])
                if iou > best_iou:
                    best_iou = iou
                    best_pred_idx = pred_idx

            record = {
                "image_name": image_name,
                "gt_class_id": gt["class_id"],
                "gt_x_center": gt["raw"][0],
                "gt_y_center": gt["raw"][1],
                "gt_width": gt["raw"][2],
                "gt_height": gt["raw"][3],
                "pred_class_id": None,
                "pred_confidence": None,
                "pred_x1": None,
                "pred_y1": None,
                "pred_x2": None,
                "pred_y2": None,
                "iou": best_iou,
            }

            if best_pred_idx != -1:
                matched_pred = pred_boxes[best_pred_idx]
                record.update(
                    {
                        "pred_class_id": matched_pred["class_id"],
                        "pred_confidence": matched_pred["confidence"],
                        "pred_x1": matched_pred["raw"][0],
                        "pred_y1": matched_pred["raw"][1],
                        "pred_x2": matched_pred["raw"][2],
                        "pred_y2": matched_pred["raw"][3],
                    }
                )
                matched_pred_indices.add(best_pred_idx)

            results.append(record)

        for pred_idx, pred in enumerate(pred_boxes):
            if pred_idx not in matched_pred_indices:
                results.append(
                    {
                        "image_name": image_name,
                        "gt_class_id": None,
                        "gt_x_center": None,
                        "gt_y_center": None,
                        "gt_width": None,
                        "gt_height": None,
                        "pred_class_id": pred["class_id"],
                        "pred_confidence": pred["confidence"],
                        "pred_x1": pred["raw"][0],
                        "pred_y1": pred["raw"][1],
                        "pred_x2": pred["raw"][2],
                        "pred_y2": pred["raw"][3],
                        "iou": 0.0,
                    }
                )

    df = pd.DataFrame(results)

    df["Match_Type"] = np.where(
        df["iou"] >= iou_threshold,
        "True Positive",
        np.where(
            df["gt_class_id"].isnull(),
            "False Positive",
            np.where(df["pred_class_id"].isnull(), "False Negative", "Low IoU (FP + FN)"),
        ),
    )

    return df


def save_to_excel(df, output_path, iou_threshold=0.5):
    total_gt = df["gt_class_id"].notna().sum()
    total_pred = df["pred_class_id"].notna().sum()
    tp = (df["Match_Type"] == "True Positive").sum()
    fp = (df["Match_Type"] == "False Positive").sum() + (df["Match_Type"] == "Low IoU (FP + FN)").sum()
    fn = (df["Match_Type"] == "False Negative").sum() + (df["Match_Type"] == "Low IoU (FP + FN)").sum()

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

    summary_data = {
        "Metric": [
            "Total Ground Truths",
            "Total Predictions",
            "IoU Threshold",
            "True Positives (TP)",
            "False Positives (FP)",
            "False Negatives (FN)",
            "Precision",
            "Recall",
            "F1-Score",
        ],
        "Value": [total_gt, total_pred, iou_threshold, tp, fp, fn, round(precision, 4), round(recall, 4), round(f1, 4)],
    }
    df_summary = pd.DataFrame(summary_data)

    df_fp = df[df["Match_Type"].isin(["False Positive", "Low IoU (FP + FN)"])].copy()
    df_fn = df[df["Match_Type"].isin(["False Negative", "Low IoU (FP + FN)"])].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="All Data", index=False)
        df_fp.to_excel(writer, sheet_name="False Positives", index=False)
        df_fn.to_excel(writer, sheet_name="False Negatives", index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=False, color="000000")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    wb.save(output_path)
    print(f"Results saved to {output_path}")


def visualize_from_dataframe(df, images_dir, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    grouped = df.groupby("image_name")

    for image_name, group in grouped:
        img_path_jpg = os.path.join(images_dir, f"{image_name}.jpg")
        img_path_png = os.path.join(images_dir, f"{image_name}.png")

        if os.path.exists(img_path_jpg):
            img_path = img_path_jpg
        elif os.path.exists(img_path_png):
            img_path = img_path_png
        else:
            continue

        img = cv2.imread(img_path)
        if img is None:
            print(f"Warning: Could not read image {img_path}")
            continue

        h, w, _ = img.shape

        for _, row in group.iterrows():
            if pd.notna(row["gt_class_id"]):
                x_c, y_c = row["gt_x_center"], row["gt_y_center"]
                bw, bh = row["gt_width"], row["gt_height"]

                x_min, y_min = x_c - bw / 2, y_c - bh / 2
                x_max, y_max = x_c + bw / 2, y_c + bh / 2

                pt1_gt = (int(x_min * w), int(y_min * h))
                pt2_gt = (int(x_max * w), int(y_max * h))

                cv2.rectangle(img, pt1_gt, pt2_gt, (0, 255, 0), 2)
                cv2.putText(
                    img,
                    f"GT: {int(row['gt_class_id'])}",
                    (pt1_gt[0], pt1_gt[1] - 5),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    (0, 255, 0),
                    2,
                )

            if pd.notna(row["pred_class_id"]):
                pt1_pred = (int(row["pred_x1"] * w), int(row["pred_y1"] * h))
                pt2_pred = (int(row["pred_x2"] * w), int(row["pred_y2"] * h))

                cv2.rectangle(img, pt1_pred, pt2_pred, (0, 0, 255), 2)

                conf = row["pred_confidence"]
                conf_str = f" {conf:.2f}" if pd.notna(conf) else ""

                iou_str = ""
                if row["iou"] > 0:
                    iou_str = f" | IoU: {row['iou']:.2f}"

                label = f"Pred: {int(row['pred_class_id'])}{conf_str}{iou_str}"

                cv2.putText(img, label, (pt1_pred[0], pt2_pred[1] + 15), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

        output_path = os.path.join(output_dir, f"{image_name}_visualized.jpg")
        cv2.imwrite(output_path, img)
    print(f"Visualizations saved to {output_dir}")


class ModelType(Enum):
    DETR = "DETR"
    YOLO = "YOLOV8S"


if __name__ == "__main__":
    MODEL_TYPE = ModelType.YOLO
    LABELS_DIR = Path(r"D:\praca_magisterska\project\Licence_plate_detection\datasets\licence_plate\labels\test")
    PREDS_DIR = Path(
        rf"D:\praca_magisterska\project\Licence_plate_detection\detection\results\{MODEL_TYPE.value}\v_001"
    )
    IMAGES_DIR = Path(r"D:\praca_magisterska\project\Licence_plate_detection\datasets\licence_plate\images\test")
    VISUALS_OUTPUT_DIR = Path(
        rf"D:\praca_magisterska\project\Licence_plate_detection\data_processing/visualization_{MODEL_TYPE.value.lower()}"
    )
    OUTPUT_EXCEL_FILE = f"{MODEL_TYPE.value.lower()}_iou_analysis_results.xlsx"
    IOU_THRESHOLD = 0.5

    df = process_dataset(LABELS_DIR, PREDS_DIR, IOU_THRESHOLD)
    save_to_excel(df, OUTPUT_EXCEL_FILE, IOU_THRESHOLD)

    visualize_from_dataframe(df, IMAGES_DIR, VISUALS_OUTPUT_DIR)
